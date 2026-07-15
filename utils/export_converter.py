import sys
import os
import csv

def convert_report(csv_path, format_choice):
    if not os.path.exists(csv_path):
        print(f"❌ Error: CSV file not found at {csv_path}")
        return

    # Check which formats to export
    export_xlsx = format_choice in ('2', '4', 'xlsx', 'xlsx-only', 'all')
    export_pdf = format_choice in ('3', '4', 'pdf', 'pdf-only', 'all')

    if not export_xlsx and not export_pdf:
        return

    # Generate output paths
    base_path = os.path.splitext(csv_path)[0]
    xlsx_path = base_path + ".xlsx"
    pdf_path = base_path + ".pdf"

    if export_xlsx:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            print(f"  → Exporting to Excel: {xlsx_path}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Audit Report"
            
            # Read CSV
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)
                
            for r_idx, row in enumerate(rows, 1):
                for c_idx, val in enumerate(row, 1):
                    # Convert to numeric if possible
                    try:
                        if '.' in val:
                            val = float(val)
                        else:
                            val = int(val)
                    except ValueError:
                        pass
                    ws.cell(row=r_idx, column=c_idx, value=val)
                    
            # Styling
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
                
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(name='Arial', size=10)
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(horizontal='left')
                        
            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            wb.save(xlsx_path)
            print(f"  ✅ Excel export successful.")
        except Exception as e:
            print(f"  ❌ Error exporting to Excel: {e}")

    if export_pdf:
        try:
            from fpdf import FPDF
            
            print(f"  → Exporting to PDF: {pdf_path}")
            
            # Read CSV data
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)
                
            if not rows:
                print("  ⚠️ CSV file is empty. Skipping PDF generation.")
                return
                
            num_cols = len(rows[0])
            orientation = 'P' if num_cols <= 5 else 'L'
            
            class PDFReport(FPDF):
                def __init__(self, title_text):
                    super().__init__()
                    self.title_text = title_text
                    
                def header(self):
                    self.set_font('helvetica', 'B', 14)
                    self.set_text_color(31, 78, 120)
                    self.cell(0, 10, self.title_text, new_x="LMARGIN", new_y="NEXT", align='L')
                    self.ln(4)
                    
                def footer(self):
                    self.set_y(-15)
                    self.set_font('helvetica', 'I', 8)
                    self.set_text_color(128, 128, 128)
                    self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', align='C')
            
            title = os.path.basename(csv_path).replace('_audit.csv', ' Audit Report').replace('_', ' ').title()
            pdf = PDFReport(title_text=title)
            pdf.alias_nb_pages()
            pdf.add_page(orientation=orientation)
            
            page_width = pdf.epw
            
            # Proportional column width calculation
            max_lens = [0] * num_cols
            for row in rows:
                for c_idx, val in enumerate(row):
                    if c_idx < num_cols:
                        max_lens[c_idx] = max(max_lens[c_idx], len(str(val)))
            
            # Ensure headers count in sizing
            for c_idx, val in enumerate(rows[0]):
                max_lens[c_idx] = max(max_lens[c_idx], len(str(val)) + 2)
                        
            total_lens = sum(max_lens) or 1
            col_widths = []
            for m_len in max_lens:
                width = (m_len / total_lens) * page_width
                col_widths.append(max(width, 15))
                
            # Scale to fit exactly
            scale = page_width / sum(col_widths)
            col_widths = [w * scale for w in col_widths]
            
            # Draw Headers
            pdf.set_font('helvetica', 'B', 9)
            pdf.set_fill_color(31, 78, 120)
            pdf.set_text_color(255, 255, 255)
            for c_idx, header in enumerate(rows[0]):
                pdf.cell(col_widths[c_idx], 8, header, border=1, align='C', fill=True)
            pdf.ln()
            
            # Draw Rows
            pdf.set_font('helvetica', '', 8)
            pdf.set_text_color(0, 0, 0)
            fill = False
            
            for row in rows[1:]:
                pdf.set_fill_color(245, 247, 250)
                for c_idx in range(num_cols):
                    val = row[c_idx] if c_idx < len(row) else ""
                    # Estimate char capacity based on width
                    max_chars = max(int(col_widths[c_idx] / 1.6), 5)
                    text = str(val)
                    if len(text) > max_chars:
                        text = text[:max_chars-3] + "..."
                        
                    align = 'R' if text.replace('.', '', 1).isdigit() else 'L'
                    pdf.cell(col_widths[c_idx], 6, text, border=1, align=align, fill=fill)
                pdf.ln()
                fill = not fill
                
            pdf.output(pdf_path)
            print(f"  ✅ PDF export successful.")
        except Exception as e:
            print(f"  ❌ Error exporting to PDF: {e}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 export_converter.py <csv_file_path> <format_choice>")
        sys.exit(1)
        
    csv_arg = sys.argv[1]
    choice_arg = sys.argv[2]
    convert_report(csv_arg, choice_arg)
